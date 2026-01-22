#!/usr/bin/env python3
"""
Re-extract column headers from SizeKorea raw files with improved header detection.

Purpose: Accurately extract real column names (Korean headers) by intelligently
detecting header rows, handling multi-row headers, and cleaning column names.

Output: Saves to verification/runs/column_inventory/<timestamp>/ (not committed).
"""

import argparse
import pandas as pd
import json
import csv
from pathlib import Path
from typing import List, Set, Dict, Optional, Tuple
from datetime import datetime
from openpyxl import load_workbook
import re


def has_korean(text: str) -> bool:
    """Check if text contains Korean characters."""
    if pd.isna(text) or not text:
        return False
    text_str = str(text)
    return any(ord(c) >= 0xAC00 and ord(c) <= 0xD7A3 for c in text_str)


def score_header_row(row_values: List, check_duplicates: bool = True) -> float:
    """
    Score a row as potential header based on multiple criteria.
    
    Args:
        row_values: List of cell values in the row
        check_duplicates: Whether to penalize duplicate values
    
    Returns:
        Score (higher is better)
    """
    if not row_values:
        return 0.0
    
    # Convert to strings and filter None/NaN
    non_empty = [str(v).strip() for v in row_values if pd.notna(v) and str(v).strip()]
    
    if not non_empty:
        return 0.0
    
    score = 0.0
    
    # 1. Unnamed/None ratio (lower is better) - weight: 40%
    unnamed_count = sum(1 for v in row_values if pd.isna(v) or str(v).strip() == '' or 
                       (isinstance(v, str) and v.startswith('Unnamed:')))
    unnamed_ratio = unnamed_count / len(row_values) if row_values else 1.0
    score += (1.0 - unnamed_ratio) * 40.0
    
    # 2. Unique non-empty strings count (higher is better) - weight: 30%
    unique_count = len(set(non_empty))
    total_count = len(row_values)
    uniqueness_ratio = unique_count / total_count if total_count > 0 else 0.0
    score += uniqueness_ratio * 30.0
    
    # 3. Korean characters presence (bonus) - weight: 20%
    korean_count = sum(1 for v in non_empty if has_korean(v))
    korean_ratio = korean_count / len(non_empty) if non_empty else 0.0
    score += korean_ratio * 20.0
    
    # 4. Duplicate penalty (lower duplicates is better) - weight: 10%
    if check_duplicates and len(non_empty) > 1:
        from collections import Counter
        counts = Counter(non_empty)
        duplicate_ratio = sum(1 for c in counts.values() if c > 1) / len(counts) if counts else 0.0
        score += (1.0 - duplicate_ratio) * 10.0
    
    return score


def find_best_header_row_xlsx(xlsx_path: Path, max_check: int = 80) -> Tuple[Optional[int], List[List], Optional[int]]:
    """
    Find best header row in Excel file by scoring multiple candidate rows.
    
    Args:
        xlsx_path: Path to Excel file
        max_check: Maximum number of rows to check
    
    Returns:
        Tuple of (best_header_row, sample_rows, potential_second_header_row)
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Read first max_check rows
        sample_rows = []
        row_scores = []
        
        for i in range(1, min(max_check + 1, ws.max_row + 1)):
            row_values = []
            for cell in ws[i]:
                row_values.append(cell.value)
            sample_rows.append(row_values)
            
            if row_values:
                score = score_header_row(row_values)
                row_scores.append((i - 1, score, row_values))  # 0-indexed
        
        wb.close()
        
        if not row_scores:
            return None, sample_rows, None
        
        # Sort by score (descending)
        row_scores.sort(key=lambda x: x[1], reverse=True)
        best_row_idx, best_score, best_row = row_scores[0]
        
        # Check if previous row might be a first header row (for multi-row headers)
        # Multi-row headers typically have: upper row (category/group) + lower row (actual names)
        potential_first = None
        if best_row_idx > 0:
            prev_row = sample_rows[best_row_idx - 1]
            prev_score = score_header_row(prev_row, check_duplicates=False)
            # If previous row also has meaningful content and looks like header (not data)
            # Check if it has codes/identifiers or category names
            prev_has_codes = any(re.match(r'^[A-Z0-9\-_\[\]]+$', str(v).strip(), re.IGNORECASE) 
                               for v in prev_row[:20] if pd.notna(v) and str(v).strip())
            if prev_score > 15.0 and (prev_has_codes or any(has_korean(str(v)) for v in prev_row[:20] if pd.notna(v))):
                # Check if it's not data (data rows usually have numeric values)
                numeric_count = sum(1 for v in prev_row[:20] if pd.notna(v) and 
                                  str(v).strip().replace('.', '').replace('-', '').isdigit())
                if numeric_count < len([v for v in prev_row[:20] if pd.notna(v)]) * 0.3:  # Less than 30% numeric
                    potential_first = best_row_idx - 1
        
        return best_row_idx, sample_rows[:min(10, len(sample_rows))], potential_first
        
    except Exception as e:
        print(f"  Warning: Error reading Excel with openpyxl: {e}")
        # Fallback to pandas
        try:
            df_sample = pd.read_excel(xlsx_path, nrows=max_check, header=None, engine='openpyxl')
            sample_rows = [df_sample.iloc[i].tolist() for i in range(min(max_check, len(df_sample)))]
            row_scores = []
            
            for i, row_values in enumerate(sample_rows):
                score = score_header_row(row_values)
                row_scores.append((i, score, row_values))
            
            if not row_scores:
                return None, sample_rows, None
            
            row_scores.sort(key=lambda x: x[1], reverse=True)
            best_row_idx, best_score, best_row = row_scores[0]
            
            potential_first = None
            if best_row_idx > 0:
                prev_row = sample_rows[best_row_idx - 1]
                prev_score = score_header_row(prev_row, check_duplicates=False)
                prev_has_codes = any(re.match(r'^[A-Z0-9\-_\[\]]+$', str(v).strip(), re.IGNORECASE) 
                                   for v in prev_row[:20] if pd.notna(v) and str(v).strip())
                if prev_score > 15.0 and (prev_has_codes or any(has_korean(str(v)) for v in prev_row[:20] if pd.notna(v))):
                    numeric_count = sum(1 for v in prev_row[:20] if pd.notna(v) and 
                                      str(v).strip().replace('.', '').replace('-', '').isdigit())
                    if numeric_count < len([v for v in prev_row[:20] if pd.notna(v)]) * 0.3:
                        potential_first = best_row_idx - 1
            
            return best_row_idx, sample_rows[:min(10, len(sample_rows))], potential_first
        except Exception as e2:
            print(f"  Error: Fallback pandas read also failed: {e2}")
            return None, [], None


def find_best_header_row_csv(csv_path: Path, encoding: str, max_check: int = 50) -> Tuple[Optional[int], List[List], Optional[int]]:
    """
    Find best header row in CSV file by scoring multiple candidate rows.
    
    Args:
        csv_path: Path to CSV file
        encoding: Encoding to use
        max_check: Maximum number of rows to check
    
    Returns:
        Tuple of (best_header_row, sample_rows, potential_second_header_row)
    """
    try:
        df_sample = pd.read_csv(csv_path, encoding=encoding, nrows=max_check, header=None, low_memory=False)
        sample_rows = [df_sample.iloc[i].tolist() for i in range(min(max_check, len(df_sample)))]
        row_scores = []
        
        for i, row_values in enumerate(sample_rows):
            score = score_header_row(row_values)
            row_scores.append((i, score, row_values))
        
        if not row_scores:
            return None, sample_rows, None
        
        row_scores.sort(key=lambda x: x[1], reverse=True)
        best_row_idx, best_score, best_row = row_scores[0]
        
        potential_first = None
        if best_row_idx > 0:
            prev_row = sample_rows[best_row_idx - 1]
            prev_score = score_header_row(prev_row, check_duplicates=False)
            prev_has_codes = any(re.match(r'^[A-Z0-9\-_\[\]]+$', str(v).strip(), re.IGNORECASE) 
                               for v in prev_row[:20] if pd.notna(v) and str(v).strip())
            if prev_score > 15.0 and (prev_has_codes or any(has_korean(str(v)) for v in prev_row[:20] if pd.notna(v))):
                numeric_count = sum(1 for v in prev_row[:20] if pd.notna(v) and 
                                  str(v).strip().replace('.', '').replace('-', '').isdigit())
                if numeric_count < len([v for v in prev_row[:20] if pd.notna(v)]) * 0.3:
                    potential_first = best_row_idx - 1
        
        return best_row_idx, sample_rows[:min(10, len(sample_rows))], potential_first
        
    except Exception as e:
        return None, [], None


def merge_multirow_headers(row1: List, row2: List) -> List[str]:
    """
    Merge two header rows into single column names.
    
    Args:
        row1: First header row (upper level)
        row2: Second header row (lower level)
    
    Returns:
        List of merged column names
    """
    merged = []
    max_len = max(len(row1), len(row2))
    
    # Check if row1 looks like codes/identifiers (mostly alphanumeric, short, no Korean)
    row1_is_code = False
    if row1:
        row1_non_empty = [str(v).strip() for v in row1 if pd.notna(v) and str(v).strip()]
        if row1_non_empty:
            code_like_count = sum(1 for v in row1_non_empty[:20] 
                                 if re.match(r'^[A-Z0-9\-_\[\]]+$', str(v).strip(), re.IGNORECASE))
            if code_like_count > len(row1_non_empty) * 0.5:  # More than 50% look like codes
                row1_is_code = True
    
    for i in range(max_len):
        val1 = str(row1[i]).strip() if i < len(row1) and pd.notna(row1[i]) else ""
        val2 = str(row2[i]).strip() if i < len(row2) and pd.notna(row2[i]) else ""
        
        # If row1 is codes/identifiers, prefer row2 (actual column names)
        if row1_is_code:
            if val2:
                merged_name = val2
            elif val1:
                merged_name = val1
            else:
                merged_name = ""
        else:
            # Normal merge: "{upper}_{lower}" or just one if the other is empty
            if val1 and val2:
                merged_name = f"{val1}_{val2}"
            elif val1:
                merged_name = val1
            elif val2:
                merged_name = val2
            else:
                merged_name = ""
        
        merged.append(merged_name)
    
    return merged


def clean_columns(columns: List[str]) -> List[str]:
    """
    Clean column names: trim, remove empty, disambiguate duplicates.
    
    Args:
        columns: List of raw column names
    
    Returns:
        List of cleaned column names
    """
    # Trim and convert to string
    cleaned = [str(col).strip() if pd.notna(col) else "" for col in columns]
    
    # Remove completely empty columns
    cleaned = [col for col in cleaned if col]
    
    # Disambiguate duplicates
    seen = {}
    result = []
    for col in cleaned:
        if col in seen:
            seen[col] += 1
            new_col = f"{col}__dup{seen[col]}"
            result.append(new_col)
        else:
            seen[col] = 0
            result.append(col)
    
    return result


def read_xlsx_headers(xlsx_path: Path) -> Dict:
    """
    Read column headers from Excel file with improved detection.
    
    Args:
        xlsx_path: Path to Excel file
    
    Returns:
        Dictionary with file info and columns
    """
    print(f"  Analyzing Excel file...")
    
    # Find best header row
    header_row, sample_rows, second_header_row = find_best_header_row_xlsx(xlsx_path, max_check=80)
    
    if header_row is None:
        return {
            "path": str(xlsx_path),
            "header_row": None,
            "n_cols": 0,
            "columns": [],
            "error": "Could not detect header row",
            "sample_rows": sample_rows[:5]
        }
    
    print(f"  Detected header row: {header_row} (0-indexed)")
    if second_header_row is not None:
        print(f"  Potential multi-row header detected (row {second_header_row} + {header_row})")
    
    try:
        # Read with detected header
        if second_header_row is not None:
            # Multi-row header: read both rows and merge (first row is upper level)
            df_row1 = pd.read_excel(xlsx_path, nrows=second_header_row+1, header=None, engine='openpyxl')
            df_row2 = pd.read_excel(xlsx_path, nrows=header_row+1, header=None, engine='openpyxl')
            
            row1_vals = df_row1.iloc[second_header_row].tolist()
            row2_vals = df_row2.iloc[header_row].tolist()
            merged_headers = merge_multirow_headers(row1_vals, row2_vals)
            
            # Read full data with second header row, then replace column names
            df = pd.read_excel(xlsx_path, header=header_row, engine='openpyxl')
            df.columns = merged_headers[:len(df.columns)]
            columns = list(df.columns)
        else:
            df = pd.read_excel(xlsx_path, header=header_row, engine='openpyxl')
            columns = list(df.columns)
        
        # Clean columns
        columns = clean_columns(columns)
        
        # Count rows
        try:
            df_full = pd.read_excel(xlsx_path, header=header_row, engine='openpyxl')
            n_rows = len(df_full)
        except:
            n_rows = None
        
        # Check for high Unnamed ratio
        unnamed_count = sum(1 for col in columns if str(col).startswith('Unnamed:') or not col)
        unnamed_ratio = unnamed_count / len(columns) if columns else 1.0
        
        if unnamed_ratio > 0.3:
            print(f"  WARNING: High Unnamed ratio ({unnamed_ratio:.1%}). Showing top 5 candidate rows:")
            for i, row in enumerate(sample_rows[:5]):
                row_preview = [str(v)[:30] for v in row[:10] if pd.notna(v)]
                print(f"    Row {i}: {row_preview}")
        
        return {
            "path": str(xlsx_path),
            "header_row": header_row,
            "n_cols": len(columns),
            "columns": columns,
            "error": None,
            "unnamed_ratio": unnamed_ratio,
            "sample_rows": sample_rows[:5] if unnamed_ratio > 0.3 else None
        }
        
    except Exception as e:
        return {
            "path": str(xlsx_path),
            "header_row": header_row,
            "n_cols": 0,
            "columns": [],
            "error": str(e),
            "sample_rows": sample_rows[:5]
        }


def read_csv_headers(csv_path: Path) -> Dict:
    """
    Read column headers from CSV file with improved detection.
    
    Args:
        csv_path: Path to CSV file
    
    Returns:
        Dictionary with file info and columns
    """
    encodings = ['utf-8-sig', 'cp949', 'euc-kr', 'utf-8']
    
    for encoding in encodings:
        try:
            print(f"  Trying encoding: {encoding}")
            
            # Find best header row
            header_row, sample_rows, second_header_row = find_best_header_row_csv(csv_path, encoding, max_check=50)
            
            if header_row is None:
                continue
            
            print(f"  Detected header row: {header_row} (0-indexed)")
            if second_header_row is not None:
                print(f"  Potential multi-row header detected (row {second_header_row} + {header_row})")
            
            # Read with detected header
            if second_header_row is not None:
                # Multi-row header (first row is upper level)
                df_row1 = pd.read_csv(csv_path, encoding=encoding, nrows=second_header_row+1, header=None, low_memory=False)
                df_row2 = pd.read_csv(csv_path, encoding=encoding, nrows=header_row+1, header=None, low_memory=False)
                
                row1_vals = df_row1.iloc[second_header_row].tolist()
                row2_vals = df_row2.iloc[header_row].tolist()
                merged_headers = merge_multirow_headers(row1_vals, row2_vals)
                
                df = pd.read_csv(csv_path, encoding=encoding, header=header_row, low_memory=False)
                df.columns = merged_headers[:len(df.columns)]
                columns = list(df.columns)
            else:
                df = pd.read_csv(csv_path, encoding=encoding, header=header_row, low_memory=False)
                columns = list(df.columns)
            
            # Clean columns
            columns = clean_columns(columns)
            
            # Count rows
            try:
                df_full = pd.read_csv(csv_path, encoding=encoding, header=header_row, low_memory=False)
                n_rows = len(df_full)
            except:
                n_rows = None
            
            # Check for high Unnamed ratio
            unnamed_count = sum(1 for col in columns if str(col).startswith('Unnamed:') or not col)
            unnamed_ratio = unnamed_count / len(columns) if columns else 1.0
            
            if unnamed_ratio > 0.3:
                print(f"  WARNING: High Unnamed ratio ({unnamed_ratio:.1%}). Showing top 5 candidate rows:")
                for i, row in enumerate(sample_rows[:5]):
                    row_preview = [str(v)[:30] for v in row[:10] if pd.notna(v)]
                    print(f"    Row {i}: {row_preview}")
            
            return {
                "path": str(csv_path),
                "header_row": header_row,
                "n_cols": len(columns),
                "columns": columns,
                "encoding": encoding,
                "error": None,
                "unnamed_ratio": unnamed_ratio,
                "sample_rows": sample_rows[:5] if unnamed_ratio > 0.3 else None
            }
            
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"  Error with encoding {encoding}: {e}")
            continue
    
    return {
        "path": str(csv_path),
        "header_row": None,
        "n_cols": 0,
        "columns": [],
        "error": "Failed to read with all attempted encodings",
        "sample_rows": []
    }


def read_file_headers(file_path: Path) -> Dict:
    """
    Read column headers from file (CSV or Excel).
    
    Args:
        file_path: Path to file
    
    Returns:
        Dictionary with file info and columns
    """
    if file_path.suffix.lower() == '.xlsx':
        return read_xlsx_headers(file_path)
    else:
        return read_csv_headers(file_path)


def create_union_csv(file_infos: Dict[str, Dict], output_path: Path):
    """
    Create CSV with union of all columns, showing which files contain each column.
    
    Args:
        file_infos: Dictionary mapping file keys to file info
        output_path: Path to save CSV
    """
    # Collect all unique columns
    all_columns = set()
    for info in file_infos.values():
        if info.get("error") is None:
            all_columns.update(info["columns"])
    
    # Create mapping: column -> which files contain it
    column_sources = {}
    for col in sorted(all_columns):
        sources = {}
        for key, info in file_infos.items():
            if info.get("error") is None and col in info["columns"]:
                sources[key] = 1
            else:
                sources[key] = 0
        column_sources[col] = sources
    
    # Write CSV
    file_keys = sorted(file_infos.keys())
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        header = ['column_name'] + [f'source_{key}' for key in file_keys]
        writer.writerow(header)
        for col in sorted(all_columns):
            sources = column_sources[col]
            row = [col] + [sources.get(key, 0) for key in file_keys]
            writer.writerow(row)


def main():
    parser = argparse.ArgumentParser(
        description="Re-extract column headers from SizeKorea raw files with improved detection"
    )
    parser.add_argument(
        "--raw_dir",
        type=str,
        default="data/raw/sizekorea_raw",
        help="Directory containing raw SizeKorea files"
    )
    parser.add_argument(
        "--out_dir",
        type=str,
        default=None,
        help="Output directory (default: verification/runs/column_inventory/<timestamp>)"
    )
    
    args = parser.parse_args()
    
    # Determine target files
    raw_dir = Path(args.raw_dir)
    
    # File mapping: key -> preferred file
    target_files = {}
    
    # 7th data (prefer xlsx, fallback to csv)
    xlsx_7th = raw_dir / "7th_data.xlsx"
    csv_7th = raw_dir / "7th_data.csv"
    if xlsx_7th.exists():
        target_files["7th"] = xlsx_7th
    elif csv_7th.exists():
        target_files["7th"] = csv_7th
    
    # 8th direct (prefer xlsx, fallback to csv)
    xlsx_8th_direct = raw_dir / "8th_data_direct.xlsx"
    csv_8th_direct = raw_dir / "8th_data_direct.csv"
    if xlsx_8th_direct.exists():
        target_files["8th_direct"] = xlsx_8th_direct
    elif csv_8th_direct.exists():
        target_files["8th_direct"] = csv_8th_direct
    
    # 8th 3d (prefer xlsx, fallback to csv)
    xlsx_8th_3d = raw_dir / "8th_data_3d.xlsx"
    csv_8th_3d = raw_dir / "8th_data_3d.csv"
    if xlsx_8th_3d.exists():
        target_files["8th_3d"] = xlsx_8th_3d
    elif csv_8th_3d.exists():
        target_files["8th_3d"] = csv_8th_3d
    
    if not target_files:
        print("Error: No target files found")
        return
    
    # Create output directory
    if args.out_dir:
        output_dir = Path(args.out_dir)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = Path(f"verification/runs/column_inventory/{timestamp}")
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("=" * 80)
    print("SizeKorea Column Header Re-extraction")
    print("=" * 80)
    print(f"\nTarget files: {len(target_files)}")
    for key, path in target_files.items():
        print(f"  {key}: {path.name}")
    print(f"\nOutput directory: {output_dir}")
    print("\n" + "=" * 80)
    
    # Read headers from each file
    file_infos = {}
    for key, file_path in target_files.items():
        print(f"\nProcessing: {file_path.name} ({key})")
        info = read_file_headers(file_path)
        file_infos[key] = info
        
        if info.get("error"):
            print(f"  ERROR: {info['error']}")
        else:
            print(f"  Header row: {info['header_row']}")
            print(f"  Columns: {info['n_cols']}")
            if 'unnamed_ratio' in info:
                print(f"  Unnamed ratio: {info['unnamed_ratio']:.1%}")
    
    # Save columns_by_file.json
    output_data = {}
    for key, info in file_infos.items():
        output_data[key] = {
            "path": info["path"],
            "header_row": info["header_row"],
            "n_cols": info["n_cols"],
            "columns": info["columns"]
        }
        if "encoding" in info:
            output_data[key]["encoding"] = info["encoding"]
    
    json_path = output_dir / "columns_by_file.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    print(f"\nSaved: {json_path}")
    
    # Create union CSV
    union_path = output_dir / "columns_union.csv"
    create_union_csv(file_infos, union_path)
    print(f"Saved: {union_path}")
    
    # Print summary
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    
    # File-by-file summary
    print("\nFile-by-file summary:")
    for key, info in file_infos.items():
        if info.get("error") is None:
            print(f"  {key}:")
            print(f"    Header row: {info['header_row']}")
            print(f"    Columns: {info['n_cols']}")
            if 'unnamed_ratio' in info:
                print(f"    Unnamed ratio: {info['unnamed_ratio']:.1%}")
    
    # Union and intersection
    column_sets = {}
    for key, info in file_infos.items():
        if info.get("error") is None:
            column_sets[key] = set(info["columns"])
    
    if column_sets:
        all_columns = set()
        for cols in column_sets.values():
            all_columns.update(cols)
        
        common_columns = set.intersection(*column_sets.values()) if len(column_sets) > 1 else column_sets[list(column_sets.keys())[0]]
        
        print(f"\nUnion columns: {len(all_columns)}")
        print(f"Intersection columns: {len(common_columns)}")
    
    print("\n" + "=" * 80)
    print(f"Output directory: {output_dir.absolute()}")
    print("=" * 80)


if __name__ == "__main__":
    main()
